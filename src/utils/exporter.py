from __future__ import annotations
import csv
import json
from pathlib import Path
from typing import Any, Dict

def export_all(result: Dict[str, Any], output_dir: str | Path = "outputs") -> Dict[str, str]:
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    files = {}
    files["json"] = str(_write_json(out / "result.json", result))
    files["markdown"] = str(_write_markdown(out / "result.md", result))
    files["storyboard_csv"] = str(_write_storyboard_csv(out / "storyboard.csv", result))
    files["prompts_csv"] = str(_write_prompts_csv(out / "prompts.csv", result))
    files["trace_json"] = str(_write_json(out / "trace.json", result.get("trace", {})))
    xlsx = _write_excel_if_available(out / "result.xlsx", result)
    if xlsx:
        files["excel"] = str(xlsx)
    return files

def _write_json(path: Path, data: Dict[str, Any]) -> Path:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return path

def _write_markdown(path: Path, result: Dict[str, Any]) -> Path:
    brief = result["product_brief"]
    eval_report = result["evaluation"]
    cost = result["cost"]
    lines = [
        f"# AI Short Video Plan: {brief['product_name']}",
        "",
        "## Product Brief",
        f"- Category: {brief.get('category')}",
        f"- Target Market: {brief.get('target_market')}",
        f"- Target Audience: {brief.get('target_audience')}",
        f"- Core Selling Points: {', '.join(brief.get('core_selling_points', []))}",
        "",
        "## Selected Storyboard",
        "| Shot | Duration | Scene | Camera | Action | Caption | Risk |",
        "|---:|---|---|---|---|---|---|",
    ]
    for shot in result["storyboard"]["storyboard"]:
        lines.append(
            f"| {shot['shot_id']} | {shot['duration']} | {shot['scene']} | {shot['camera']} | {shot['action']} | {shot['caption']} | {shot['ai_risk']} |"
        )
    lines += [
        "",
        "## Quality Score",
        f"- Overall Score: {eval_report['overall_score']}",
        f"- Pass Status: {eval_report['pass_status']}",
        "",
        "## Token & Cost Estimate",
        f"- Agent Count: {cost['agent_count']}",
        f"- Input Tokens: {cost['input_tokens']}",
        f"- Output Tokens: {cost['output_tokens']}",
        f"- Total Tokens: {cost['total_tokens']}",
        f"- Estimated Cost USD: {cost['estimated_cost_usd']}",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")
    return path

def _write_storyboard_csv(path: Path, result: Dict[str, Any]) -> Path:
    rows = result["storyboard"]["storyboard"]
    fields = ["shot_id", "duration", "scene", "camera", "action", "product_focus", "caption", "ai_risk", "generation_note"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return path

def _write_prompts_csv(path: Path, result: Dict[str, Any]) -> Path:
    rows = result["prompts"]["prompts"]
    fields = ["model", "shot_id", "positive_prompt", "negative_prompt", "aspect_ratio", "duration", "notes"]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return path

def _write_excel_if_available(path: Path, result: Dict[str, Any]) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    title_fill = PatternFill("solid", fgColor="EAF2F8")
    header_fill = PatternFill("solid", fgColor="F4F6F7")
    thin = Side(style="thin", color="D5D8DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    brief = result["product_brief"]
    rows = [
        ["AI Short Video Factory Agent", ""],
        ["Product", brief.get("product_name")],
        ["Category", brief.get("category")],
        ["Market", brief.get("target_market")],
        ["Audience", brief.get("target_audience")],
        ["Overall Score", result["evaluation"].get("overall_score")],
        ["Total Tokens", result["cost"].get("total_tokens")],
        ["Estimated Cost USD", result["cost"].get("estimated_cost_usd")],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:B1")
    for row in ws.iter_rows(min_row=2, max_row=len(rows), max_col=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        row[0].font = Font(bold=True)
        row[0].fill = header_fill
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80

    ws2 = wb.create_sheet("Storyboard")
    headers = ["Shot", "Duration", "Scene", "Camera", "Action", "Focus", "Caption", "Risk"]
    ws2.append(headers)
    for h in ws2[1]:
        h.font = Font(bold=True)
        h.fill = header_fill
        h.border = border
    for shot in result["storyboard"]["storyboard"]:
        ws2.append([shot["shot_id"], shot["duration"], shot["scene"], shot["camera"], shot["action"], shot["product_focus"], shot["caption"], shot["ai_risk"]])
    for col, width in zip("ABCDEFGH", [8, 12, 36, 36, 44, 34, 38, 12]):
        ws2.column_dimensions[col].width = width
    for row in ws2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    ws3 = wb.create_sheet("Prompts")
    headers = ["Model", "Shot", "Positive Prompt", "Negative Prompt", "Duration", "Notes"]
    ws3.append(headers)
    for h in ws3[1]:
        h.font = Font(bold=True)
        h.fill = header_fill
        h.border = border
    for p in result["prompts"]["prompts"]:
        ws3.append([p["model"], p["shot_id"], p["positive_prompt"], p["negative_prompt"], p["duration"], p["notes"]])
    for col, width in zip("ABCDEF", [12, 8, 80, 54, 12, 38]):
        ws3.column_dimensions[col].width = width
    for row in ws3.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    wb.save(path)
    return path
